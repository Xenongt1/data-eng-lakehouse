#!/usr/bin/env python3
"""Split monthly xlsx files into per-day CSVs.

Reads Data/orders_apr_2025.xlsx and Data/order_items_apr_2025.xlsx, where each
sheet name is an ISO date (e.g. ``2025-04-01``), and writes one CSV per sheet
to Data/daily_csvs/ named ``orders_<date>.csv`` / ``order_items_<date>.csv``.
Also copies Data/products.csv into Data/daily_csvs/ for symmetry.

The multi-sheet xlsx is a delivery convenience; the lakehouse pipeline ingests
one CSV at a time, so this script simulates 15 daily file drops.
"""
from __future__ import annotations

import csv
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "Data"
OUT = DATA / "daily_csvs"

SOURCES = [
    ("orders", DATA / "orders_apr_2025.xlsx"),
    ("order_items", DATA / "order_items_apr_2025.xlsx"),
]


def split_workbook(prefix: str, xlsx_path: Path, out_dir: Path) -> int:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    written = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            print(f"  skip empty sheet {sheet_name}", file=sys.stderr)
            continue
        out_path = out_dir / f"{prefix}_{sheet_name}.csv"
        with out_path.open("w", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(header)
            row_count = 0
            for row in rows:
                writer.writerow(row)
                row_count += 1
        print(f"  wrote {out_path.name} ({row_count} rows)")
        written += 1
    wb.close()
    return written


def main() -> int:
    if not DATA.exists():
        print(f"Data directory not found: {DATA}", file=sys.stderr)
        return 1
    OUT.mkdir(parents=True, exist_ok=True)

    total = 0
    for prefix, xlsx in SOURCES:
        if not xlsx.exists():
            print(f"missing source file: {xlsx}", file=sys.stderr)
            return 1
        print(f"splitting {xlsx.name}")
        total += split_workbook(prefix, xlsx, OUT)

    products_src = DATA / "products.csv"
    if products_src.exists():
        shutil.copy2(products_src, OUT / "products.csv")
        print(f"  copied {products_src.name}")
        total += 1
    else:
        print(f"warning: {products_src} not found", file=sys.stderr)

    print(f"done: {total} files in {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
